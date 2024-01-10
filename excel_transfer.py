import pandas as pd
from openpyxl import Workbook, load_workbook, styles
from openpyxl.styles import Alignment

#input restaurant data into spreadsheet
def inputDataIntoSpreadsheet(path, data, row):
      
      if path == None:
            #create new workbook, keep in mind path
            print("ERROR, SPREADSHEET NOT CREATED")
      
      workbook = load_workbook(filename = path)
      sheet = workbook.active

      #equation in case more data is needed
      ##currLetter = 'A'
      #for v in data:
            #sheet[str(currLetter) + str(row)] = v
            #currLetter = chr(ord(currLetter) + 1)
      
      sheet["A" + str(row)] = data["name"]
      sheet["B" + str(row)] = data["description"]
      sheet["C" + str(row)] = data["address"]
      sheet["D" + str(row)] = data["maplink"]
      
      workbook.save(path)
      return path

#setup workbook

#TODO:
#1. Format City text to make sure its right
#. Format Titles so they look nice

def createSpreadsheet(city):
      workbook = Workbook()
      sheet = workbook.active
     
     #Create Title
      sheet.merge_cells("A1:E1")
      sheet["A1"] = "" + city + " Restaurant Data"
      sheet["A1"].alignment = Alignment(horizontal = "center", vertical = "center")

      #set data titles
      sheet["A2"] = "Restaurant Name :"
      sheet["B2"] = "Description: "
      sheet["C2"] = "Address:"
      sheet["D2"] = "Google Maps:"

      #Save workbook with city name
      path = "" + city + "_restaurant_data.xlsx"
      workbook.save(path)

      #return path to workbook with format, add checkers for error
      return path